# -*- coding: utf-8 -*-
from datetime import timedelta, datetime
from odoo import api, fields, models, _
import io
from odoo.tools import config, DEFAULT_SERVER_DATE_FORMAT, DEFAULT_SERVER_DATETIME_FORMAT, parse_version

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    import xlrd
    try:
        from xlrd import xlsx
    except ImportError:
        xlsx = None
except ImportError:
    xlrd = xlsx = None

class Import(models.TransientModel):
    _inherit = 'base_import.import'

    def _read_xlsx(self, options):
        if xlsx:
            return self._read_xls(options)

        import openpyxl.cell.cell as types
        import openpyxl.styles.numbers as styles  # noqa: PLC0415
        import base64 as _b64  # local alias
        import zipfile as _zip
        import xml.etree.ElementTree as _ET
        book = load_workbook(io.BytesIO(self.file or b''), data_only=True)
        sheets = options['sheets'] = book.sheetnames
        sheet_name = options['sheet'] = options.get('sheet') or sheets[0]
        sheet = book[sheet_name]
        # Build image map: (row_index, col_index) -> base64 string
        image_map = {}
        try:
            for img in getattr(sheet, "_images", []) or []:
                row = col = None
                try:
                    marker = getattr(img, 'anchor', None)
                    marker_from = getattr(marker, '_from', None) or getattr(marker, 'from', None)
                    if marker_from is not None and hasattr(marker_from, 'col') and hasattr(marker_from, 'row'):
                        col = int(marker_from.col + 1)
                        row = int(marker_from.row + 1)
                except Exception:
                    pass
                if (row is None or col is None) and isinstance(getattr(img, 'anchor', None), str):
                    try:
                        from openpyxl.utils import coordinate_to_tuple as _coord
                        r_idx, c_idx = _coord(img.anchor)
                        row, col = int(r_idx), int(c_idx)
                    except Exception:
                        row = col = None

                blob = None
                try:
                    data_attr = getattr(img, '_data', None)
                    if callable(data_attr):
                        raw = data_attr()
                        if hasattr(raw, 'getvalue'):
                            blob = raw.getvalue()
                        elif isinstance(raw, (bytes, bytearray)):
                            blob = bytes(raw)
                    elif isinstance(data_attr, (bytes, bytearray)):
                        blob = bytes(data_attr)
                    if blob is None and hasattr(img, 'image') and img.image is not None:
                        try:
                            bio = io.BytesIO()
                            img.image.save(bio, format=getattr(img.image, 'format', None) or 'PNG')
                            blob = bio.getvalue()
                        except Exception:
                            pass
                except Exception:
                    blob = None

                if blob and row and col:
                    image_map[(row, col)] = _b64.b64encode(blob).decode()
        except Exception:
            image_map = {}

        if not image_map and (self.file or b''):
            try:
                zf = _zip.ZipFile(io.BytesIO(self.file))
                sheet_index = book.sheetnames.index(sheet.title)
                sheet_path = f"xl/worksheets/sheet{sheet_index + 1}.xml"
                rels_path = f"xl/worksheets/_rels/sheet{sheet_index + 1}.xml.rels"
                if sheet_path in zf.namelist():
                    sheet_xml = _ET.fromstring(zf.read(sheet_path))
                    ns = {
                        'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
                    }
                    drawing_elems = sheet_xml.findall('.//a:drawing', ns)
                    drawing_rids = [e.attrib.get('{%s}id' % ns['r']) for e in drawing_elems if e is not None]
                    drawing_files = []
                    if rels_path in zf.namelist():
                        rels_xml = _ET.fromstring(zf.read(rels_path))
                        for rel in rels_xml.findall(
                                './/{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                            rid = rel.attrib.get('Id')
                            target = rel.attrib.get('Target', '')
                            if rid in drawing_rids and target:
                                if target.startswith('../drawings/'):
                                    drawing_files.append('xl/drawings/' + target.split('../drawings/')[1])
                                elif target.startswith('drawings/'):
                                    drawing_files.append('xl/' + target)

                    d_ns = {
                        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
                    }
                    for dfile in drawing_files:
                        if dfile not in zf.namelist():
                            continue
                        dxml = _ET.fromstring(zf.read(dfile))
                        d_rels_file = dfile.replace('drawings/', 'drawings/_rels/') + '.rels'
                        embed_to_media = {}
                        if d_rels_file in zf.namelist():
                            drels_xml = _ET.fromstring(zf.read(d_rels_file))
                            for rel in drels_xml.findall(
                                    './/{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                if rel.attrib.get('Type', '').endswith('/image'):
                                    embed_to_media[rel.attrib.get('Id')] = rel.attrib.get('Target', '')

                        for anchor_tag in ('oneCellAnchor', 'twoCellAnchor'):
                            for anc in dxml.findall(f'.//xdr:{anchor_tag}', d_ns):
                                from_node = anc.find('.//xdr:from', d_ns)
                                if from_node is None:
                                    continue
                                col_node = from_node.find('xdr:col', d_ns)
                                row_node = from_node.find('xdr:row', d_ns)
                                if col_node is None or row_node is None:
                                    continue
                                col = int((col_node.text or '0')) + 1
                                row = int((row_node.text or '0')) + 1

                                blip = anc.find('.//a:blip', d_ns)
                                if blip is None:
                                    continue
                                embed = blip.attrib.get('{%s}embed' % d_ns['r'])
                                media_rel = embed_to_media.get(embed)
                                if not media_rel:
                                    continue

                                if media_rel.startswith('../'):
                                    media_rel = media_rel.replace('../', 'xl/')
                                if media_rel.startswith('media/'):
                                    media_rel = 'xl/' + media_rel
                                if media_rel.startswith('drawings/'):
                                    candidate = media_rel.replace('drawings/', 'media/')
                                    if candidate not in zf.namelist():
                                        base_name = media_rel.split('/')[-1]
                                        for name in zf.namelist():
                                            if name.startswith('xl/media/') and name.split('/')[-1] == base_name:
                                                candidate = name
                                                break
                                    media_rel = candidate

                                if media_rel in zf.namelist():
                                    blob = zf.read(media_rel)
                                    if blob:
                                        image_map[(row, col)] = _b64.b64encode(blob).decode()
            except Exception:
                pass
        rows = []
        for rowx, row in enumerate(sheet.rows, 1):
            values = []
            for colx, cell in enumerate(row, 1):
                if cell.data_type is types.TYPE_ERROR:
                    raise ValueError(
                        _("Invalid cell value at row %(row)s, column %(col)s: %(cell_value)s", row=rowx, col=colx,
                          cell_value=cell.value)
                    )

                if cell.value is None:
                    # inject base64 image if present at this cell
                    if (rowx, colx) in image_map:
                        values.append(image_map[(rowx, colx)])
                    else:
                        values.append('')
                elif isinstance(cell.value, float):
                    if cell.value % 1 == 0:
                        values.append(str(int(cell.value)))
                    else:
                        values.append(str(cell.value))
                elif cell.is_date:
                    d_fmt = styles.is_datetime(cell.number_format)
                    if d_fmt == "datetime":
                        values.append(cell.value.strftime(DEFAULT_SERVER_DATETIME_FORMAT))
                    elif d_fmt == "date":
                        values.append(cell.value.strftime(DEFAULT_SERVER_DATE_FORMAT))
                    else:
                        raise ValueError(
                            _("Invalid cell format at row %(row)s, column %(col)s: %(cell_value)s, with format: %(cell_format)s, as (%(format_type)s) formats are not supported.",
                              row=rowx, col=colx, cell_value=cell.value, cell_format=cell.number_format,
                              format_type=d_fmt)
                        )
                else:
                    # if cell looks empty string but has image, inject base64
                    if (isinstance(cell.value, str) and not cell.value.strip()) and (rowx, colx) in image_map:
                        values.append(image_map[(rowx, colx)])
                    else:
                        values.append(str(cell.value))

            if any(x.strip() for x in values):
                rows.append(values)
        return sheet.max_row, rows